import os
from openpyxl import load_workbook
import openpyxl

ficheiro_excel = 'C:/Users/eudis/OneDrive/Documentos/Fotos Patio 2024/Pasta 5 - Copiar.xlsx'
diretório = './'

wb_obj = openpyxl.load_workbook(ficheiro_excel)
sheet_obj = wb_obj.active
qtd = sheet_obj.max_row

for i in range(2, qtd + 1):
    Descrição = sheet_obj.cell(row = i, column = 1).value
    Placa = sheet_obj.cell(row = i, column = 2).value
    Valor = sheet_obj.cell(row = i, column = 3).value
    Sucata = sheet_obj.cell(row = i, column = 4).value

   # if not os.path.exists(diretório+Descrição.upper()+'_'+Placa+'_'+Valor+'_'+Sucata):
    if not os.path.exists(diretório+Descrição.upper()+'_'+Placa+'_'+Valor):
       os.makedirs(diretório+Descrição.upper()+'_'+Placa+'_'+Valor)
