import os
from openpyxl import load_workbook

# Caminho para o arquivo Excel
excel_file = 'C:/Users/eudis/OneDrive/Documentos/Fotos Patio 2024/Pasta 5 - mesclado.xlsx'

# Abra o arquivo Excel
workbook = load_workbook(excel_file)

# Seleciona a aba desejada (por exemplo, a primeira aba)
worksheet = workbook.active

# Lista de colunas que contém os nomes das pastas
colunas_com_nomes_de_pasta = [1,2,3,4]
# Iterar por cada coluna que contém nomes de pastas
for coluna_idx in colunas_com_nomes_de_pasta:
    # Percorra as linhas da coluna (ignora a primeira linha se for o cabeçalho)
    for row in worksheet.iter_rows(min_col=coluna_idx, max_col=coluna_idx, values_only=True):
        # Obtenha o nome da pasta da célula atual
        folder_name = row[0]
        
        # Verifica se o nome da pasta não é vazio
        if folder_name:
            # Cria a pasta se ela não existir
            os.makedirs(folder_name, exist_ok=True)
